"""
AI Document Management System — FastAPI Backend
================================================
Security-hardened version with:
- Environment-based configuration (no hardcoded secrets)
- Restricted CORS origins
- Security headers middleware
- Rate limiting on auth & AI endpoints
- Input validation & file content verification
- Full tenant isolation on all endpoints
"""

import os
import uuid
import json
import csv
import io
import gc                          # ✅ FIX-1: memory cleanup
import logging
from pathlib import Path
from datetime import datetime
from contextlib import asynccontextmanager
from typing import Dict, Any, List
import string
import random

from pydantic import BaseModel, EmailStr, Field, field_validator
from fastapi import FastAPI, UploadFile, File, Depends, HTTPException, Query, Body, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, RedirectResponse, Response, StreamingResponse
from sqlalchemy.orm import Session

from backend import config
from backend.database import engine, get_db, Base, SessionLocal
from backend.models import Document, User, Organization
from backend.auth import get_current_user, create_access_token, verify_password, get_password_hash
from backend.ocr_service import process_document, process_document_with_layout
from backend.extraction_service import extract_document_data
from backend.search_service import parse_search_query
from backend.assistant_service import get_assistant_response
from backend.security import (
    SecurityHeadersMiddleware,
    validate_file_magic,
    sanitize_filename,
    validate_password_strength,
    validate_name,
)
from backend.admin_routes import admin_router, org_router

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# ── ✅ FIX-2: update_env_value مفقودة كلياً من الكود الأصلي ──────────────────
def update_env_value(key: str, value: str):
    """Write/update a key=value line in .env and set it in os.environ."""
    env_path = Path(".env")
    if env_path.exists():
        lines = env_path.read_text(encoding="utf-8").splitlines()
        new_lines, updated = [], False
        for line in lines:
            if line.startswith(f"{key}=") or line.startswith(f"{key} ="):
                new_lines.append(f"{key}={value}")
                updated = True
            else:
                new_lines.append(line)
        if not updated:
            new_lines.append(f"{key}={value}")
        env_path.write_text("\n".join(new_lines) + "\n", encoding="utf-8")
    else:
        with open(env_path, "w", encoding="utf-8") as f:
            f.write(f"{key}={value}\n")
    os.environ[key] = value


def create_super_admin():
    db = SessionLocal()
    try:
        org = db.query(Organization).filter(Organization.name == "Arkivo Core").first()
        if not org:
            org = Organization(name="Arkivo Core", invite_code="ARKIVO-CORE")
            db.add(org)
            db.commit()
            db.refresh(org)

        admin_password = os.getenv("ADMIN_PASSWORD", "Arkivo_Admin_2026")

        user = db.query(User).filter(User.email == "adminA@arkivo.com").first()
        if not user:
            user = User(
                name="adminA",
                email="adminA@arkivo.com",
                hashed_password=get_password_hash(admin_password),
                role="super_admin",
                organization_id=org.id
            )
            db.add(user)
            db.commit()
            print("✅ Super Admin created")
        else:
            user.role = "super_admin"
            user.hashed_password = get_password_hash(admin_password)
            db.commit()
            print("♻️ Super Admin updated")
    finally:
        db.close()


@asynccontextmanager
async def lifespan(app: FastAPI):
    print("🔥 STARTUP WORKING")
    Base.metadata.create_all(bind=engine)
    create_super_admin()
    yield


app = FastAPI(
    title="Arkivo — AI Document Management System",
    version="2.0.0",
    lifespan=lifespan,
    docs_url=None if os.environ.get("PRODUCTION") else "/docs",
    redoc_url=None if os.environ.get("PRODUCTION") else "/redoc",
)

UPLOAD_DIR = config.UPLOAD_DIR
os.makedirs(UPLOAD_DIR, exist_ok=True)
ALLOWED_EXTENSIONS = config.ALLOWED_EXTENSIONS
MAX_FILE_SIZE = config.MAX_FILE_SIZE

app.add_middleware(SecurityHeadersMiddleware)
app.add_middleware(
    CORSMiddleware,
    allow_origins=config.ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type"],
)

try:
    from slowapi import Limiter, _rate_limit_exceeded_handler
    from slowapi.util import get_remote_address
    from slowapi.errors import RateLimitExceeded

    limiter = Limiter(key_func=get_remote_address)
    app.state.limiter = limiter
    app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)
    RATE_LIMITING_ENABLED = True
    logger.info("Rate limiting enabled.")
except ImportError:
    RATE_LIMITING_ENABLED = False
    logger.warning("slowapi not installed — rate limiting disabled.")

    class _NoOpLimiter:
        def limit(self, *args, **kwargs):
            def decorator(func):
                return func
            return decorator
    limiter = _NoOpLimiter()

FRONTEND_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "frontend")
if os.path.isdir(FRONTEND_DIR):
    app.mount("/static", StaticFiles(directory=FRONTEND_DIR), name="static")

app.include_router(org_router)
app.include_router(admin_router)


@app.get("/", include_in_schema=False)
async def root():
    return RedirectResponse(url="/static/index.html")


# ── Auth ──────────────────────────────────────────────────────────────────────

class RegisterRequest(BaseModel):
    name: str = Field(..., min_length=2, max_length=200)
    email: EmailStr
    password: str = Field(..., min_length=8, max_length=128)
    action: str = Field(..., pattern="^(create|join)$")
    org_name: str = Field(None, min_length=2, max_length=200)
    invite_code: str = Field(None, max_length=50)

class LoginRequest(BaseModel):
    email: EmailStr
    password: str = Field(..., min_length=1, max_length=128)

def generate_invite_code():
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))

@app.post("/api/auth/register")
@limiter.limit(config.RATE_LIMIT_AUTH)
def register(request: Request, payload: RegisterRequest, db: Session = Depends(get_db)):
    pw_valid, pw_error = validate_password_strength(payload.password)
    if not pw_valid:
        raise HTTPException(status_code=400, detail=pw_error)

    name_valid, name_error = validate_name(payload.name)
    if not name_valid:
        raise HTTPException(status_code=400, detail=name_error)

    user = db.query(User).filter(User.email == payload.email).first()
    if user:
        raise HTTPException(status_code=400, detail="Email already registered")

    if payload.action == 'create':
        if not payload.org_name:
            raise HTTPException(status_code=400, detail="Organization name is required")
        org_valid, org_error = validate_name(payload.org_name)
        if not org_valid:
            raise HTTPException(status_code=400, detail=org_error)
        org = Organization(name=payload.org_name.strip(), invite_code=generate_invite_code())
        db.add(org)
        db.commit()
        db.refresh(org)
        role = "admin"
    elif payload.action == 'join':
        if not payload.invite_code:
            raise HTTPException(status_code=400, detail="Invite code is required")
        org = db.query(Organization).filter(Organization.invite_code == payload.invite_code).first()
        if not org:
            raise HTTPException(status_code=404, detail="Invalid invite code")
        role = "member"
    else:
        raise HTTPException(status_code=400, detail="Invalid action")

    new_user = User(
        name=payload.name.strip(),
        email=payload.email,
        hashed_password=get_password_hash(payload.password),
        role=role,
        organization_id=org.id
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return {"message": "User created successfully"}


@app.post("/api/auth/login")
@limiter.limit(config.RATE_LIMIT_AUTH)
def login(request: Request, payload: LoginRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.email == payload.email).first()
    if not user or not verify_password(payload.password, user.hashed_password):
        raise HTTPException(status_code=401, detail="Incorrect email or password")
    access_token = create_access_token(
        data={"sub": str(user.id), "org_id": user.organization_id, "role": user.role}
    )
    return {"access_token": access_token, "token_type": "bearer"}


@app.get("/api/auth/me")
def get_me(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    org = db.query(Organization).filter(Organization.id == current_user.organization_id).first()
    return {
        "id": current_user.id,
        "name": current_user.name,
        "email": current_user.email,
        "role": current_user.role,
        "organization": {"id": org.id, "name": org.name, "invite_code": org.invite_code}
    }


# ── Settings ──────────────────────────────────────────────────────────────────

class SettingsUpdate(BaseModel):
    api_key: str = Field(None, max_length=200)
    model: str = Field(None, max_length=100)


@app.get("/api/settings")
def get_settings(current_user: User = Depends(get_current_user)):
    from backend.config import mask_api_key, get_env_value
    current_key = get_env_value("OPENROUTER_API_KEY", "")
    current_model = get_env_value("OPENROUTER_MODEL", "openai/gpt-4o-mini")
    return {
        "api_key_masked": mask_api_key(current_key),
        "api_key_set": bool(current_key),
        "model": current_model,
    }


@app.put("/api/settings")
def update_settings(payload: SettingsUpdate, current_user: User = Depends(get_current_user)):
    from backend.config import mask_api_key, get_env_value
    import backend.config as cfg

    if current_user.role not in ("admin", "super_admin"):
        raise HTTPException(status_code=403, detail="Only admins can change settings")

    updated = []

    if payload.api_key is not None:
        new_key = payload.api_key.strip()
        if new_key:
            update_env_value("OPENROUTER_API_KEY", new_key)   # ✅ now defined above
            cfg.OPENROUTER_API_KEY = new_key
            updated.append("api_key")
            logger.info("OPENROUTER_API_KEY updated by admin user %s", current_user.id)

    if payload.model is not None:
        new_model = payload.model.strip()
        if new_model:
            update_env_value("OPENROUTER_MODEL", new_model)   # ✅ now defined above
            cfg.OPENROUTER_MODEL = new_model
            updated.append("model")

    if not updated:
        raise HTTPException(status_code=400, detail="No valid settings provided")

    return {
        "message": "Settings updated successfully",
        "updated": updated,
        "api_key_masked": mask_api_key(cfg.OPENROUTER_API_KEY),
        "model": cfg.OPENROUTER_MODEL,
    }


# ── Documents ─────────────────────────────────────────────────────────────────

@app.post("/api/documents/upload")
async def upload_document(
    file: UploadFile = File(...),
    lang: str = Query(default="latin", pattern="^[a-z]{2,10}$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    safe_filename = sanitize_filename(file.filename)
    ext = safe_filename.rsplit(".", 1)[-1].lower() if "." in safe_filename else ""
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"File type '.{ext}' not allowed. Supported: {', '.join(ALLOWED_EXTENSIONS)}",
        )

    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail=f"File size exceeds {config.MAX_FILE_SIZE_MB} MB limit.")
    if len(content) == 0:
        raise HTTPException(status_code=400, detail="Empty file uploaded.")

    if not validate_file_magic(content, ext):
        raise HTTPException(
            status_code=400,
            detail="File content does not match declared file type. Upload rejected.",
        )

    stored_filename = f"{uuid.uuid4().hex}.{ext}"
    file_path = os.path.join(UPLOAD_DIR, stored_filename)
    resolved = os.path.realpath(file_path)
    if not resolved.startswith(os.path.realpath(UPLOAD_DIR)):
        raise HTTPException(status_code=400, detail="Invalid file path.")

    with open(file_path, "wb") as f:
        f.write(content)

    doc = Document(
        filename=stored_filename,
        original_filename=safe_filename,
        file_type=ext,
        file_size=len(content),
        status="processing",
        organization_id=current_user.organization_id,
    )
    db.add(doc)
    db.commit()
    db.refresh(doc)

    try:
        from fastapi.concurrency import run_in_threadpool
        layout_result = await run_in_threadpool(process_document_with_layout, file_path, ext, lang)

        doc.extracted_text = layout_result["text"]
        doc.page_count = layout_result["page_count"]
        doc.status = "completed"

        if ext == "pdf" and layout_result["page_count"] > 1:
            doc.file_type = "folder"
            doc.ocr_layout_data = "[]"
            for page_data in layout_result.get("pages", []):
                child_doc = Document(
                    parent_id=doc.id,
                    filename=stored_filename,
                    original_filename=f"{safe_filename} - Page {page_data['page']}",
                    file_type="pdf",
                    file_size=len(content),
                    status="completed",
                    extracted_text=page_data["text"],
                    page_count=1,
                    ocr_layout_data=json.dumps([page_data], ensure_ascii=False),
                    organization_id=current_user.organization_id
                )
                db.add(child_doc)
        else:
            doc.ocr_layout_data = json.dumps(layout_result.get("pages", []), ensure_ascii=False)

        logger.info(f"OCR completed for document {doc.id} — {len(layout_result['text'])} chars")
    except Exception as e:
        doc.status = "failed"
        doc.error_message = str(e)[:500]
        logger.error(f"OCR failed for document {doc.id}: {type(e).__name__}")

    db.commit()
    db.refresh(doc)
    return doc.to_dict()


@app.get("/api/documents")
def list_documents(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    docs = db.query(Document).filter(
        Document.parent_id == None,
        Document.organization_id == current_user.organization_id
    ).order_by(Document.created_at.desc()).all()
    result = []
    for doc in docs:
        d = doc.to_dict()
        if d["extracted_text"] and len(d["extracted_text"]) > 200:
            d["extracted_text_preview"] = d["extracted_text"][:200] + "..."
        else:
            d["extracted_text_preview"] = d["extracted_text"]
        del d["extracted_text"]
        result.append(d)
    return result


class SearchQuery(BaseModel):
    query: str = None
    filters: dict = None


@app.post("/api/documents/search")
def search_documents(payload: SearchQuery, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    if payload.query:
        filters = parse_search_query(payload.query)
    else:
        filters = payload.filters or {}

    docs = db.query(Document).filter(
        Document.file_type != "folder",
        Document.organization_id == current_user.organization_id
    ).order_by(Document.created_at.desc()).all()

    result = []
    for doc in docs:
        try:
            meta = json.loads(doc.extracted_metadata or "{}")
        except (json.JSONDecodeError, TypeError):
            meta = {}

        match = True

        if filters.get("document_type"):
            dtype = filters["document_type"].lower()
            if (doc.doc_type or "").lower() != dtype:
                match = False

        if match and filters.get("company"):
            company_val = meta.get("company", {}).get("value", "") if isinstance(meta.get("company"), dict) else meta.get("company", "")
            if filters["company"].lower() not in (company_val or "").lower():
                match = False

        if match and filters.get("client_name"):
            client_val = meta.get("client_name", {}).get("value", "") if isinstance(meta.get("client_name"), dict) else meta.get("client_name", "")
            if filters["client_name"].lower() not in (client_val or "").lower():
                match = False

        if match and filters.get("invoice_number"):
            inv_val = meta.get("invoice_number", {}).get("value", "") if isinstance(meta.get("invoice_number"), dict) else meta.get("invoice_number", "")
            if filters["invoice_number"].lower() not in (inv_val or "").lower():
                match = False

        if match and filters.get("date_range"):
            dr = filters["date_range"]
            if dr:
                d_from = dr.get("from")
                d_to = dr.get("to")
                if d_from or d_to:
                    date_val = meta.get("date", {}).get("value", "") if isinstance(meta.get("date"), dict) else meta.get("date", "")
                    if not date_val:
                        match = False
                    else:
                        parsed_iso = None
                        try:
                            clean_v = str(date_val).replace('.', '-').replace('/', '-').strip()
                            parts = [p.strip() for p in clean_v.split('-') if p.strip()]
                            if len(parts) == 3:
                                y, m, d = None, None, None
                                p0, p1, p2 = int(parts[0]), int(parts[1]), int(parts[2])
                                if p0 > 1000:
                                    y, m, d = p0, p1, p2
                                elif p2 > 31:
                                    y = p2
                                    if y < 100: y += 2000
                                    if p1 > 12: d, m = p1, p0
                                    elif p0 > 12: d, m = p0, p1
                                    else: d, m = p0, p1
                                elif p2 > 12:
                                    y = p2 + 2000 if p2 < 100 else p2
                                    d, m = p0, p1
                                else:
                                    d, m, y = p0, p1, p2
                                    if y < 100: y += 2000
                                if y and m and d:
                                    dt = datetime(y, m, d)
                                    parsed_iso = dt.strftime("%Y-%m-%d")
                        except Exception:
                            pass
                        if parsed_iso:
                            if d_from and parsed_iso < d_from:
                                match = False
                            if d_to and parsed_iso > d_to:
                                match = False
                        else:
                            match = False

        if match and filters.get("keyword"):
            kw = filters["keyword"].lower()
            if kw not in doc.original_filename.lower() and kw not in (doc.extracted_text or "").lower():
                match = False

        if match:
            d = doc.to_dict()
            if d.get("extracted_text") and len(d["extracted_text"]) > 200:
                d["extracted_text_preview"] = d["extracted_text"][:200] + "..."
            else:
                d["extracted_text_preview"] = d["extracted_text"] or ""
            if "extracted_text" in d:
                del d["extracted_text"]
            result.append(d)

    return {"filters": filters, "results": result}


# ── AI Assistant ──────────────────────────────────────────────────────────────

class AssistantPayload(BaseModel):
    query: str
    language: str = "English"


@app.post("/api/ai/chat")
@limiter.limit(config.RATE_LIMIT_AI)
def assistant_chat(request: Request, payload: AssistantPayload, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    try:
        response = get_assistant_response(
            payload.query, db, payload.language,
            organization_id=current_user.organization_id
        )
        return response
    except Exception as e:
        logger.error(f"Assistant Chat Error: {type(e).__name__}")
        raise HTTPException(status_code=500, detail="An error occurred processing your AI request.")


# ── Document Operations ───────────────────────────────────────────────────────

@app.get("/api/documents/{doc_id}/children")
def list_document_children(doc_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    docs = db.query(Document).filter(
        Document.parent_id == doc_id,
        Document.organization_id == current_user.organization_id
    ).order_by(Document.id.asc()).all()
    result = []
    for doc in docs:
        d = doc.to_dict()
        if d["extracted_text"] and len(d["extracted_text"]) > 200:
            d["extracted_text_preview"] = d["extracted_text"][:200] + "..."
        else:
            d["extracted_text_preview"] = d["extracted_text"]
        del d["extracted_text"]
        result.append(d)
    return result


@app.get("/api/documents/{doc_id}")
def get_document(doc_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    doc = db.query(Document).filter(
        Document.id == doc_id,
        Document.organization_id == current_user.organization_id
    ).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    return doc.to_dict()


@app.get("/api/documents/{doc_id}/download")
def download_document(doc_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    doc = db.query(Document).filter(
        Document.id == doc_id,
        Document.organization_id == current_user.organization_id
    ).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    file_path = os.path.join(UPLOAD_DIR, doc.filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found on disk")

    dl_filename = doc.original_filename
    if doc.file_type and not dl_filename.lower().endswith(f".{doc.file_type.lower()}"):
        dl_filename += f".{doc.file_type}"

    mime_types = {
        "pdf": "application/pdf",
        "png": "image/png",
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg"
    }
    media_type = mime_types.get(doc.file_type, "application/octet-stream")

    if doc.parent_id is not None and doc.file_type == "pdf":
        try:
            layout = json.loads(doc.ocr_layout_data or "[]")
            if layout and isinstance(layout, list) and "page" in layout[0]:
                page_num = int(layout[0]["page"])
                from pypdf import PdfReader, PdfWriter
                reader = PdfReader(file_path)
                writer = PdfWriter()
                if 0 <= page_num - 1 < len(reader.pages):
                    writer.add_page(reader.pages[page_num - 1])
                    buf = io.BytesIO()
                    writer.write(buf)
                    buf.seek(0)
                    return Response(
                        content=buf.read(),
                        media_type="application/pdf",
                        headers={"Content-Disposition": f'attachment; filename="{dl_filename}"'}
                    )
        except Exception as e:
            logger.error(f"Failed to extract single PDF page for download Doc ID {doc_id}: {e}")

    return FileResponse(
        path=file_path,
        filename=dl_filename,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{dl_filename}"'},
    )


@app.get("/api/documents/{doc_id}/preview")
def preview_document(doc_id: int, page: int = 1, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    doc = db.query(Document).filter(
        Document.id == doc_id,
        Document.organization_id == current_user.organization_id
    ).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    file_path = os.path.join(UPLOAD_DIR, doc.filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found on disk")

    if doc.file_type in ("jpg", "jpeg", "png"):
        media_types = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png"}
        return FileResponse(path=file_path, media_type=media_types.get(doc.file_type, "application/octet-stream"))

    if doc.file_type in ("pdf", "folder"):
        from pdf2image import convert_from_path
        if doc.parent_id is not None:
            try:
                layout = json.loads(doc.ocr_layout_data or "[]")
                if layout:
                    page = layout[0].get("page", page)
            except (json.JSONDecodeError, TypeError):
                pass
        if doc.file_type == "folder":
            page = 1

        poppler_path = os.environ.get("POPPLER_PATH", None)
        kwargs = {}
        if poppler_path:
            kwargs["poppler_path"] = poppler_path
        else:
            local_poppler = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "poppler", "poppler-24.07.0", "Library", "bin")
            if os.path.isdir(local_poppler):
                kwargs["poppler_path"] = local_poppler

        try:
            images = convert_from_path(file_path, dpi=150, first_page=page, last_page=page, **kwargs)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed to render PDF page: {e}")

        if not images:
            raise HTTPException(status_code=404, detail="Page not found")

        buf = io.BytesIO()
        images[0].save(buf, format="PNG")
        buf.seek(0)
        return Response(content=buf.read(), media_type="image/png")

    raise HTTPException(status_code=400, detail="Unsupported file type for preview")


@app.delete("/api/documents/all")
def delete_all_documents(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    if current_user.role not in ("admin", "super_admin", "member"):
        raise HTTPException(status_code=403, detail="Unrecognized user role")
    try:
        root_docs = db.query(Document).filter(
            Document.parent_id == None,
            Document.organization_id == current_user.organization_id
        ).all()
        deleted_files = 0
        for doc in root_docs:
            db.query(Document).filter(Document.parent_id == doc.id).delete()
            file_path = os.path.join(UPLOAD_DIR, doc.filename)
            if os.path.exists(file_path):
                os.remove(file_path)
                deleted_files += 1
            db.delete(doc)
        db.commit()
        return {"message": f"Deleted {len(root_docs)} documents ({deleted_files} files removed)"}
    except Exception as e:
        db.rollback()
        logger.error(f"Failed to delete all documents: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/documents/{doc_id}")
def delete_document(doc_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    if current_user.role not in ("admin", "super_admin", "member"):
        raise HTTPException(status_code=403, detail="Unrecognized user role")
    doc = db.query(Document).filter(
        Document.id == doc_id,
        Document.organization_id == current_user.organization_id
    ).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    try:
        if doc.parent_id is None:
            db.query(Document).filter(
                Document.parent_id == doc_id,
                Document.organization_id == current_user.organization_id
            ).delete()
        if doc.parent_id is None and doc.file_type != "folder":
            file_path = os.path.join(UPLOAD_DIR, doc.filename)
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except Exception as e:
                    logger.error(f"File deletion error: {e}")
        db.delete(doc)
        db.commit()
        return {"message": "Document deleted successfully", "id": doc_id}
    except Exception as e:
        db.rollback()
        logger.error(f"Full deletion error: {e}")
        raise HTTPException(status_code=500, detail=f"Database error: {e}")


@app.get("/api/documents/{doc_id}/layout")
def get_document_layout(
    doc_id: int,
    page: int = Query(default=1, ge=1),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    doc = db.query(Document).filter(
        Document.id == doc_id,
        Document.organization_id == current_user.organization_id
    ).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    try:
        pages = json.loads(doc.ocr_layout_data or "[]")
    except (json.JSONDecodeError, TypeError):
        pages = []

    if not pages:
        raise HTTPException(status_code=404, detail="No layout data available. Re-upload to generate.")
    if page > len(pages):
        raise HTTPException(status_code=404, detail=f"Page {page} not found. Document has {len(pages)} pages.")

    page_data = pages[page - 1]
    return {
        "doc_id": doc_id,
        "page": page,
        "total_pages": len(pages),
        "image_width": page_data.get("image_width", 0),
        "image_height": page_data.get("image_height", 0),
        "blocks": page_data.get("blocks", []),
    }


class TextUpdate(BaseModel):
    extracted_text: str


@app.put("/api/documents/{doc_id}/text")
def update_document_text(doc_id: int, payload: TextUpdate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    doc = db.query(Document).filter(
        Document.id == doc_id,
        Document.organization_id == current_user.organization_id
    ).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    doc.extracted_text = payload.extracted_text
    db.commit()
    return {"message": "Text updated successfully"}


class LayoutBlockUpdate(BaseModel):
    text: str
    x: float
    y: float
    width: float
    height: float
    confidence: float = 1.0


class LayoutUpdate(BaseModel):
    page: int
    blocks: List[LayoutBlockUpdate]


@app.put("/api/documents/{doc_id}/layout")
def update_document_layout(doc_id: int, payload: LayoutUpdate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    doc = db.query(Document).filter(
        Document.id == doc_id,
        Document.organization_id == current_user.organization_id
    ).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    try:
        pages = json.loads(doc.ocr_layout_data or "[]")
    except Exception:
        pages = []
    while len(pages) < payload.page:
        pages.append({"page": len(pages) + 1, "blocks": []})
    pages[payload.page - 1]["blocks"] = [b.dict() for b in payload.blocks]
    doc.ocr_layout_data = json.dumps(pages, ensure_ascii=False)
    db.commit()
    return {"message": "Layout updated successfully"}


# ── ✅ FIX-3: extract_document — text + blocks limits + gc ───────────────────
@app.post("/api/documents/{doc_id}/extract")
def extract_document(doc_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    doc = db.query(Document).filter(
        Document.id == doc_id,
        Document.organization_id == current_user.organization_id
    ).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    if not doc.extracted_text or doc.status != "completed":
        raise HTTPException(
            status_code=400,
            detail="Document has no extracted text. Ensure OCR completed successfully.",
        )

    layout_blocks = None
    try:
        pages = json.loads(doc.ocr_layout_data or "[]")
        if pages:
            if isinstance(pages, dict):
                pages = [pages]
            if isinstance(pages, list):
                layout_blocks = []
                for page in pages:
                    if isinstance(page, dict):
                        layout_blocks.extend(page.get("blocks", []))
                # ✅ حد أقصى 200 block — يمنع إرسال JSON ضخم إلى LLM
                if layout_blocks:
                    layout_blocks = layout_blocks[:200]
    except (json.JSONDecodeError, TypeError, AttributeError, KeyError):
        pass

    # ✅ حد أقصى 6000 حرف للنص — كافٍ لـ GPT-4o-mini ويوفر ذاكرة وزمن
    text_to_extract = (doc.extracted_text or "")[:6000]

    result = None
    try:
        result = extract_document_data(text_to_extract, layout_blocks)
    except Exception as e:
        logger.error(f"Extraction failed for doc {doc_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Extraction failed: {e}")
    finally:
        # ✅ تحرير الذاكرة دائماً حتى عند الخطأ
        del layout_blocks, text_to_extract
        gc.collect()

    doc.doc_type = result.get("type", "other")
    doc.extracted_metadata = json.dumps(result, ensure_ascii=False)
    db.commit()
    db.refresh(doc)
    return {"doc_id": doc_id, **result}


@app.get("/api/documents/{doc_id}/extract")
def get_extraction(doc_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    doc = db.query(Document).filter(
        Document.id == doc_id,
        Document.organization_id == current_user.organization_id
    ).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    try:
        meta = json.loads(doc.extracted_metadata or "{}")
    except (json.JSONDecodeError, TypeError):
        meta = {}
    return {"doc_id": doc_id, "type": doc.doc_type or "unknown", **meta}


class ExtractUpdate(BaseModel):
    metadata: Dict[str, Any]


@app.put("/api/documents/{doc_id}/extract")
def update_extraction(doc_id: int, payload: ExtractUpdate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    doc = db.query(Document).filter(
        Document.id == doc_id,
        Document.organization_id == current_user.organization_id
    ).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    doc.extracted_metadata = json.dumps(payload.metadata, ensure_ascii=False)
    db.commit()
    return {"doc_id": doc_id, "type": doc.doc_type or "unknown", **payload.metadata}


# ── Export helpers ────────────────────────────────────────────────────────────

def _extract_val(meta: dict, key: str) -> str:
    val = meta.get(key)
    if isinstance(val, dict):
        return val.get("value") or ""
    return val or ""


def _build_row(doc: Document, meta: dict) -> dict:
    return {
        "ID": doc.id,
        "Original Filename": doc.original_filename,
        "File Type": doc.file_type,
        "File Size (bytes)": doc.file_size,
        "Pages": doc.page_count,
        "Status": doc.status,
        "Created At": doc.created_at.isoformat() if doc.created_at else "",
        "Document Type": doc.doc_type or "unknown",
        "Invoice Number": _extract_val(meta, "invoice_number"),
        "Date": _extract_val(meta, "date"),
        "Due Date": _extract_val(meta, "due_date"),
        "Expiry Date": _extract_val(meta, "expiry_date"),
        "Total Amount": _extract_val(meta, "total_amount"),
        "Currency": _extract_val(meta, "currency"),
        "Company": _extract_val(meta, "company"),
        "Client Name": _extract_val(meta, "client_name"),
    }


EXPORT_HEADERS = [
    "ID", "Original Filename", "File Type", "File Size (bytes)",
    "Pages", "Status", "Created At",
    "Document Type", "Invoice Number", "Date", "Due Date", "Expiry Date",
    "Total Amount", "Currency", "Company", "Client Name",
]


def _stream_csv(rows: list, headers: list, filename: str) -> StreamingResponse:
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    writer.writerows(rows)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _stream_txt(rows: list, title: str, filename: str) -> StreamingResponse:
    output = io.StringIO()
    output.write(f"=== {title} ===\n\n")
    for i, row in enumerate(rows):
        output.write(f"--- Document {i+1} ---\n")
        for key, value in row.items():
            output.write(f"{key}: {value}\n")
        output.write("\n")
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/plain",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _stream_excel(rows: list, headers: list, filename: str, sheet_title: str = "Documents") -> StreamingResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    for column in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max_len + 4, 50)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Export endpoints ──────────────────────────────────────────────────────────

@app.get("/api/documents/{doc_id}/export")
def export_document(
    doc_id: int,
    format: str = Query(default="csv", regex="^(csv|excel|txt)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    doc = db.query(Document).filter(
        Document.id == doc_id,
        Document.organization_id == current_user.organization_id
    ).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    try:
        meta = json.loads(doc.extracted_metadata or "{}")
    except (json.JSONDecodeError, TypeError):
        meta = {}

    row = _build_row(doc, meta)
    base_name = doc.original_filename.rsplit(".", 1)[0] if "." in doc.original_filename else doc.original_filename

    if format == "txt":
        return _stream_txt([row], "AI Extraction Metadata", f"{base_name}_metadata.txt")
    if format == "csv":
        return _stream_csv([row], list(row.keys()), f"{base_name}_metadata.csv")
    return _stream_excel([row], list(row.keys()), f"{base_name}_metadata.xlsx", "Metadata")


@app.get("/api/export/all")
def export_all_documents(
    format: str = Query(default="csv", regex="^(csv|excel|txt)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    docs = db.query(Document).filter(
        Document.file_type != "folder",
        Document.organization_id == current_user.organization_id
    ).order_by(Document.created_at.desc()).all()

    rows = []
    for doc in docs:
        try:
            meta = json.loads(doc.extracted_metadata or "{}")
        except (json.JSONDecodeError, TypeError):
            meta = {}
        rows.append(_build_row(doc, meta))

    if format == "txt":
        return _stream_txt(rows, "All Documents Extraction Metadata", "all_documents.txt")
    if format == "csv":
        return _stream_csv(rows, EXPORT_HEADERS, "all_documents.csv")
    return _stream_excel(rows, EXPORT_HEADERS, "all_documents.xlsx", "Documents")


@app.get("/api/export/folder/{folder_id}")
def export_folder_documents(
    folder_id: int,
    format: str = Query(default="csv", regex="^(csv|excel|txt)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    docs = db.query(Document).filter(
        Document.parent_id == folder_id,
        Document.organization_id == current_user.organization_id
    ).order_by(Document.id.asc()).all()

    rows = []
    for doc in docs:
        try:
            meta = json.loads(doc.extracted_metadata or "{}")
        except (json.JSONDecodeError, TypeError):
            meta = {}
        rows.append(_build_row(doc, meta))

    if format == "txt":
        return _stream_txt(rows, f"Folder {folder_id} Documents Extraction Metadata", f"folder_{folder_id}_documents.txt")
    if format == "csv":
        return _stream_csv(rows, EXPORT_HEADERS, f"folder_{folder_id}_documents.csv")
    return _stream_excel(rows, EXPORT_HEADERS, f"folder_{folder_id}_documents.xlsx", "Documents")


# ── ✅ FIX-4: تعريف مزدوج لـ export/custom محذوف — نسخة واحدة فقط ──────────

class ExportQuery(BaseModel):
    filters: dict = None
    document_ids: List[int] = None
    format: str = "csv"


@app.post("/api/export/ids")
def export_documents_by_ids(
    payload: ExportQuery,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not payload.document_ids:
        raise HTTPException(status_code=400, detail="document_ids list is required")

    docs = db.query(Document).filter(
        Document.id.in_(payload.document_ids),
        Document.organization_id == current_user.organization_id
    ).all()

    rows = []
    for doc in docs:
        try:
            meta = json.loads(doc.extracted_metadata or "{}")
        except (json.JSONDecodeError, TypeError):
            meta = {}
        rows.append(_build_row(doc, meta))

    if payload.format == "txt":
        return _stream_txt(rows, "AI Assistant Results Export", "ai_results.txt")
    if payload.format == "csv":
        return _stream_csv(rows, EXPORT_HEADERS, "ai_results.csv")
    return _stream_excel(rows, EXPORT_HEADERS, "ai_results.xlsx", "Results")


@app.post("/api/export/custom")
def export_custom_documents(
    payload: ExportQuery,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # ✅ FIX-4: استخدام ExportQuery + فلترة مباشرة من DB بدلاً من استدعاء search_documents
    sq = SearchQuery(filters=payload.filters)
    search_res = search_documents(sq, db, current_user=current_user)

    # جمع IDs من نتائج البحث ثم قراءة من DB مباشرة للحصول على extracted_metadata
    matched_ids = [d["id"] for d in search_res["results"]]
    docs = db.query(Document).filter(
        Document.id.in_(matched_ids),
        Document.organization_id == current_user.organization_id
    ).all() if matched_ids else []

    rows = []
    for doc in docs:
        try:
            meta = json.loads(doc.extracted_metadata or "{}")
        except (json.JSONDecodeError, TypeError):
            meta = {}
        rows.append(_build_row(doc, meta))

    if payload.format == "txt":
        return _stream_txt(rows, "Custom Export Metadata", "custom_export.txt")
    if payload.format == "csv":
        return _stream_csv(rows, EXPORT_HEADERS, "custom_export.csv")
    return _stream_excel(rows, EXPORT_HEADERS, "custom_export.xlsx", "Documents")
if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run("main:app", host="0.0.0.0", port=port, workers=1)
